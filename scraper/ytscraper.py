from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import openpyxl
import time

# Load the workbook
workbook = openpyxl.load_workbook('G:/Python/PROJEKT/scraper/allvideos.xlsx')
sheet = workbook.active

# Assuming links are in column B starting from the first row (including header)
column_index = 2
starting_row = 1

links = [sheet.cell(row=i, column=column_index).value for i in range(starting_row + 1, sheet.max_row + 1)]

# Scrape descriptions
descriptions = []

# Configure the browser
firefox_binary_path = r"C:\Program Files\Mozilla Firefox\firefox.exe"
options = webdriver.FirefoxOptions()
options.binary_location = firefox_binary_path
driver = webdriver.Firefox(options=options)

# Add a new column header if it doesn't exist
new_column_index = sheet.max_column + 1
if sheet.cell(row=starting_row, column=new_column_index).value is None:
    sheet.cell(row=starting_row, column=new_column_index, value='Description from XPath')

for i, link in enumerate(links):
    try:
        driver.get(link)

        # Accept cookies if a dialog is present
        try:
            cookies_accept_button = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,
                                                "/html/body/ytd-app/ytd-consent-bump-v2-lightbox/tp-yt-paper-dialog/div[4]/div[2]/div[6]/div[1]/ytd-button-renderer[2]/yt-button-shape/button/yt-touch-feedback-shape/div/div[2]"))
            )
            cookies_accept_button.click()
        except:
            pass

        # Click "Show More" button if available
        try:
            show_more_button = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='expand']"))
            )
            show_more_button.click()
        except:
            pass

        # Scroll to the description box if available
        try:
            description_box = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//div[@id='description']"))
            )
            driver.execute_script("arguments[0].scrollIntoView();", description_box)
            WebDriverWait(driver, 10).until(
                EC.text_to_be_present_in_element((By.XPATH, "//div[@id='description']//*[normalize-space()]"),
                                                 "Last updated on")
            )
        except:
            pass

        # Scrape the description
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        description = soup.find('div', {'id': 'description'})
        if description:
            description_text = description.get_text(strip=True)
            descriptions.append(description_text)
            sheet.cell(row=i + starting_row + 1, column=new_column_index, value=description_text)
        else:
            descriptions.append('Description not found')
            sheet.cell(row=i + starting_row + 1, column=new_column_index, value='Description not found')

        # Save the changes after each description
        workbook.save('G:/Python/PROJEKT/scraper/allvideos_with_descriptions.xlsx')

    except Exception as e:
        print(f"Error for URL: {link}")
        print(e)
        descriptions.append('Error')
        sheet.cell(row=i + starting_row + 1, column=new_column_index, value='Error')
        workbook.save('G:/Python/PROJEKT/scraper/allvideos_with_descriptions.xlsx')

# Close the browser
driver.quit()
